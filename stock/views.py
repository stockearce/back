import requests
import base64
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation

# Django Core & Shortcuts
from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.contrib import messages

# Authentication & Decorators
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User

# Database & Querying
from django.db.models import F, Sum, Q

# External Libraries
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Local Models
from .models import (
    Producto, TipoProducto, Cliente, Ventas,
    DetalleVenta, ImagenProducto, Chofer
)


# ==================================
# AUTH
# ==================================
def login_usuario(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            messages.success(request, f'Bienvenido {user.username}')
            return redirect('home')
        else:
            messages.error(request, 'Usuario o contraseña incorrectos')

    return render(request, 'auth/login.html')


@login_required
def logout_usuario(request):
    logout(request)
    messages.success(request, 'Sesión cerrada correctamente')
    return redirect('login')


# ==================================
# HOME
# ==================================
@login_required
def home(request):
    productos_bajo_stock = Producto.objects.filter(
        cantidad__lt=F('umbral_alerta')
    ).count()

    con_alerta = productos_bajo_stock > 0

    ventas_pendientes = Ventas.objects.filter(
        estado__in=['pendiente', 'confirmada']
    ).count()

    contexto = {
        'productos_bajo_stock': productos_bajo_stock,
        'con_alerta': con_alerta,
        'ventas_pendientes': ventas_pendientes,
    }

    return render(request, 'home.html', contexto)


# ==================================
# PRODUCTOS
# ==================================
from django.core.paginator import Paginator
@login_required
def lista_productos(request):
    tipos        = TipoProducto.objects.all()
    query_nombre = request.GET.get('nombre')
    query_tipo   = request.GET.get('tipo')

    productos = Producto.objects.select_related('tipo').all()

    if query_nombre:
        productos = productos.filter(nombre__icontains=query_nombre)
    if query_tipo:
        productos = productos.filter(tipo_id=query_tipo)

    # Productos con stock bajo (se muestran arriba de la tabla)
    productos_alerta = Producto.objects.filter(
        cantidad__lt=F('umbral_alerta')
    ).select_related('tipo').order_by('nombre')

    # Paginación — 50 por página
    from django.core.paginator import Paginator
    paginator = Paginator(productos, 50)
    page      = request.GET.get('page', 1)
    productos = paginator.get_page(page)

    return render(request, 'lista_productos.html', {
        'productos':        productos,
        'tipos':            tipos,
        'query_nombre':     query_nombre,
        'query_tipo':       query_tipo,
        'productos_alerta': productos_alerta,
    })
@login_required
def crear_tipo(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        if nombre:
            TipoProducto.objects.create(nombre=nombre)
            return redirect('lista_productos')
    return render(request, 'crear_tipo.html')


@login_required
def crear_producto(request):
    tipos = TipoProducto.objects.all()

    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        tipo_id = request.POST.get('tipo')
        cantidad = request.POST.get('cantidad')
        valor_compra = request.POST.get('valor_compra')
        valor = request.POST.get('valor')
        umbral_alerta = request.POST.get('umbral_alerta', 5)

        try:
            umbral_alerta = int(umbral_alerta)
        except (ValueError, TypeError):
            umbral_alerta = 5

        if nombre and tipo_id:
            tipo = TipoProducto.objects.get(id=tipo_id)
            Producto.objects.create(
                nombre=nombre,
                tipo=tipo,
                cantidad=cantidad,
                valor_compra=valor_compra,
                valor=valor,
                umbral_alerta=umbral_alerta
            )
            return redirect('lista_productos')

    return render(request, 'crear_producto.html', {'tipos': tipos})


@login_required
def actualizar_stock(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)

    if request.method == 'POST':
        cantidad = int(request.POST.get('cantidad'))
        accion = request.POST.get('accion')

        if accion == 'sumar':
            producto.cantidad += cantidad
        elif accion == 'restar':
            producto.cantidad -= cantidad
            if producto.cantidad < 0:
                producto.cantidad = 0

        producto.save()
        return redirect('lista_productos')

    return render(request, 'actualizar_stock.html', {'producto': producto})


@login_required
def editar_producto(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)
    tipos = TipoProducto.objects.all()

    if request.method == 'POST':
        nombre        = request.POST.get('nombre', '').strip()
        tipo_id       = request.POST.get('tipo')
        valor         = request.POST.get('valor', '0')
        valor_compra  = request.POST.get('valor_compra', '0')
        cantidad      = request.POST.get('cantidad', '0')
        umbral_alerta = request.POST.get('umbral_alerta', '5')

        producto.nombre  = nombre
        producto.tipo_id = tipo_id

        try:
            producto.valor = Decimal(valor)
        except InvalidOperation:
            messages.error(request, 'El precio de venta no es válido.')
            return redirect('editar_producto', producto_id=producto.id)

        try:
            producto.valor_compra = Decimal(valor_compra)
        except InvalidOperation:
            messages.error(request, 'El precio de compra no es válido.')
            return redirect('editar_producto', producto_id=producto.id)

        try:
            producto.cantidad = int(cantidad)
        except ValueError:
            messages.error(request, 'La cantidad debe ser un número entero.')
            return redirect('editar_producto', producto_id=producto.id)

        try:
            producto.umbral_alerta = int(umbral_alerta)
        except ValueError:
            messages.error(request, 'El umbral debe ser un número entero.')
            return redirect('editar_producto', producto_id=producto.id)

        producto.save()
        messages.success(request, f'Producto "{producto.nombre}" actualizado.')
        return redirect('lista_productos')

    return render(request, 'editar_producto.html', {
        'producto': producto,
        'tipos': tipos,
    })


@login_required
def panel_alertas(request):
    productos_con_alerta = Producto.objects.filter(
        cantidad__lt=F('umbral_alerta')
    ).select_related('tipo').order_by('nombre')

    contexto = {
        'productos_con_alerta': productos_con_alerta,
        'conteo_alertas': productos_con_alerta.count(),
    }

    return render(request, 'panel_alertas.html', contexto)


# ==================================
# CLIENTES
# ==================================
@login_required
def lista_clientes(request):
    if request.user.is_superuser:
        clientes = Cliente.objects.all()
    else:
        perfil = request.user.perfilusuario

        if perfil.tipo == 'ventas':
            clientes = Cliente.objects.filter(usuario=request.user)
        elif perfil.tipo == 'administrativo':
            clientes = Cliente.objects.all()
        else:
            clientes = Cliente.objects.none()

    return render(request, 'lista_clientes.html', {'clientes': clientes})


@login_required
def crear_cliente(request):
    if request.method == 'POST':
        nombre_completo = request.POST.get('nombre_completo')
        nombre_local    = request.POST.get('nombre_local')
        cuil            = request.POST.get('cuil')
        email           = request.POST.get('email')
        telefono        = request.POST.get('telefono')
        direccion       = request.POST.get('direccion')

        if nombre_completo:
            Cliente.objects.create(
                usuario=request.user,
                nombre_completo=nombre_completo,
                nombre_local=nombre_local if nombre_local else None,
                cuil=cuil if cuil else None,
                email=email if email else None,
                telefono=telefono if telefono else None,
                direccion=direccion if direccion else None
            )
            messages.success(request, 'Cliente creado correctamente')
            return redirect('lista_clientes')

    return render(request, 'crear_cliente.html')


@login_required
def editar_cliente(request, cliente_id):
    cliente = get_object_or_404(Cliente, id=cliente_id)
    perfil  = getattr(request.user, 'perfilusuario', None)

    if not request.user.is_superuser and perfil.tipo == 'ventas':
        if cliente.usuario != request.user:
            messages.error(request, "No tenés permiso para editar este cliente.")
            return redirect('lista_clientes')

    vendedores = None
    if request.user.is_superuser or perfil.tipo == 'administrativo':
        vendedores = User.objects.filter(perfilusuario__tipo='ventas')

    if request.method == 'POST':
        cliente.nombre_completo = request.POST.get('nombre_completo')
        cliente.nombre_local    = request.POST.get('nombre_local')
        cliente.email           = request.POST.get('email') or None
        cliente.telefono        = request.POST.get('telefono') or None
        cliente.direccion       = request.POST.get('direccion') or None

        if request.user.is_superuser or perfil.tipo == 'administrativo':
            nuevo_usuario_id = request.POST.get('usuario')
            if nuevo_usuario_id:
                cliente.usuario = User.objects.get(id=nuevo_usuario_id)

        cliente.save()
        messages.success(request, "Cliente actualizado correctamente.")
        return redirect('lista_clientes')

    return render(request, 'editar_cliente.html', {
        'cliente': cliente,
        'vendedores': vendedores
    })


# ==================================
# VENTAS
# ==================================


@login_required
def crear_venta(request):
    if request.method == 'POST':
        cliente_id    = request.POST.get('cliente')
        notas         = request.POST.get('notas')
        productos_ids = request.POST.getlist('productos')
        cantidades    = request.POST.getlist('cantidades')

        if cliente_id and productos_ids:
            try:
                cliente = Cliente.objects.get(id=cliente_id)

                venta = Ventas.objects.create(
                    cliente=cliente,
                    estado='pendiente',
                    notas=notas or '',
                    valor_total=0,
                    usuario_creador=request.user
                )

                total_venta = 0

                for prod_id, cant in zip(productos_ids, cantidades):
                    if prod_id and cant:
                        producto = Producto.objects.get(id=prod_id)
                        cantidad = int(cant)
                        precio   = producto.valor
                        subtotal = precio * cantidad

                        DetalleVenta.objects.create(
                            venta=venta,
                            producto=producto,
                            cantidad=cantidad,
                            precio_unitario=precio,
                            subtotal=subtotal
                        )

                        total_venta += subtotal

                venta.valor_total = total_venta
                venta.save(update_fields=['valor_total'])

                messages.success(
                    request,
                    f'✅ Venta #{venta.id} creada por ${total_venta:.2f}. Estado: Pendiente.'
                )
                return redirect('lista_ventas')

            except Exception as e:
                messages.error(request, f'Error al crear venta: {str(e)}')
        else:
            messages.error(request, 'Seleccione un cliente y al menos un producto')

    productos = Producto.objects.all().order_by('nombre')

    return render(request, 'ventas/crear_venta.html', {
        'productos': productos,
    })


@login_required
def buscar_clientes(request):
    q = request.GET.get('q', '').strip()
    if len(q) < 1:
        return JsonResponse({'resultados': []})

    if request.user.is_superuser or request.user.perfilusuario.tipo == 'administrativo':
        clientes = Cliente.objects.filter(
            nombre_completo__icontains=q
        ).values('id', 'nombre_completo', 'telefono')[:10]
    else:
        clientes = Cliente.objects.filter(
            nombre_completo__icontains=q,
            usuario=request.user
        ).values('id', 'nombre_completo', 'telefono')[:10]

    resultados = [
        {
            'id': c['id'],
            'nombre': c['nombre_completo'],
            'detalle': c['telefono'] or ''
        }
        for c in clientes
    ]
    return JsonResponse({'resultados': resultados})

@login_required
def buscar_productos(request):
    q = request.GET.get('q', '').strip()
    if len(q) < 1:
        return JsonResponse({'resultados': []})

    productos = Producto.objects.filter(
        nombre__icontains=q
    ).values('id', 'nombre', 'valor', 'cantidad')[:10]

    resultados = [
        {
            'id': p['id'],
            'nombre': p['nombre'],
            'precio': float(p['valor']),
            'stock': p['cantidad']
        }
        for p in productos
    ]
    return JsonResponse({'resultados': resultados})


@login_required
def lista_ventas(request):
    estado = request.GET.get('estado')
    fecha  = request.GET.get('fecha')

    ventas     = Ventas.objects.none()
    hay_filtros = any([estado, fecha])

    if hay_filtros:
        # Superusuario y administrativo ven todas las ventas
        if request.user.is_superuser or request.user.perfilusuario.tipo == 'administrativo':
            ventas = Ventas.objects.all()
        else:
            ventas = Ventas.objects.filter(usuario_creador=request.user)

        if fecha:
            ventas = ventas.filter(fecha_creacion__date=fecha)
        if estado:
            ventas = ventas.filter(estado=estado)

        ventas = ventas.order_by('-fecha_creacion')

    return render(request, 'ventas/lista_ventas.html', {
        'ventas':      ventas,
        'estado':      estado,
        'fecha':       fecha,
        'estados':     Ventas.ESTADO_CHOICES,
        'hay_filtros': hay_filtros,
    })


@login_required
def detalle_venta(request, venta_id):
    venta    = get_object_or_404(Ventas, id=venta_id)
    detalles = venta.detalles.select_related('producto').all()
    choferes = Chofer.objects.filter(activo=True).order_by('nombre_completo')

    return render(request, 'ventas/detalle_venta.html', {
        'venta': venta,
        'detalles': detalles,
        'choferes': choferes
    })


@login_required
def actualizar_estado_venta(request, venta_id):
    venta = get_object_or_404(Ventas, id=venta_id)

    if request.method == 'POST':
        nuevo_estado = request.POST.get('estado')

        transiciones_permitidas = {
            'pendiente':  ['confirmada', 'cancelada'],
            'confirmada': ['pendiente', 'cancelada', 'enviada'],
            'enviada':    ['entregada'],
            'entregada':  [],
            'cancelada':  []
        }

        if nuevo_estado not in transiciones_permitidas.get(venta.estado, []):
            messages.error(
                request,
                f'❌ No se puede cambiar de "{venta.get_estado_display()}" a '
                f'"{dict(Ventas.ESTADO_CHOICES).get(nuevo_estado)}"'
            )
            return redirect('detalle_venta', venta_id=venta_id)

        # Cancelar venta confirmada → devolver stock
        if nuevo_estado == 'cancelada' and venta.estado == 'confirmada':
            try:
                for detalle in venta.detalles.select_related('producto').all():
                    detalle.producto.cantidad = F('cantidad') + detalle.cantidad
                    detalle.producto.save()
                    detalle.producto.refresh_from_db()

                venta.estado = nuevo_estado
                venta.save()
                messages.success(request, '✅ Venta cancelada y stock devuelto')
            except Exception as e:
                messages.error(request, f'❌ Error al devolver stock: {str(e)}')

        # Volver a pendiente desde confirmada → devolver stock
        elif nuevo_estado == 'pendiente' and venta.estado == 'confirmada':
            try:
                for detalle in venta.detalles.select_related('producto').all():
                    detalle.producto.cantidad = F('cantidad') + detalle.cantidad
                    detalle.producto.save()
                    detalle.producto.refresh_from_db()

                venta.estado  = nuevo_estado
                venta.chofer  = None
                venta.save()
                messages.success(request, '✅ Venta vuelta a pendiente y stock devuelto')
            except Exception as e:
                messages.error(request, f'❌ Error al devolver stock: {str(e)}')

        else:
            venta.estado = nuevo_estado
            venta.save()
            messages.success(request, f'✅ Estado actualizado: {venta.get_estado_display()}')

        return redirect('detalle_venta', venta_id=venta_id)

    return render(request, 'ventas/actualizar_estado.html', {
        'venta': venta,
        'estados': Ventas.ESTADO_CHOICES
    })


# ==================================
# CONSULTAR VENTAS (REPORTES)
# ==================================
@login_required
def consultar_ventas(request):
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    vendedor_id = request.GET.get('vendedor')

    ventas = Ventas.objects.none()

    if fecha_desde and fecha_hasta:
        fecha_desde_dt = datetime.strptime(fecha_desde, "%Y-%m-%d")
        fecha_hasta_dt = datetime.strptime(fecha_hasta, "%Y-%m-%d") + timedelta(days=1)

        ventas = Ventas.objects.select_related(
            'cliente', 'chofer', 'usuario_creador'
        ).prefetch_related(
            'detalles__producto'
        ).filter(
            estado='entregada',
            fecha_envio__gte=fecha_desde_dt,
            fecha_envio__lt=fecha_hasta_dt
        )

        if vendedor_id:
            ventas = ventas.filter(usuario_creador_id=vendedor_id)

        ventas = ventas.order_by('-fecha_envio')

    if request.GET.get('exportar') == 'excel':
        return exportar_ventas_excel(ventas, fecha_desde, fecha_hasta)

    vendedores = User.objects.filter(
        is_active=True,
        perfilusuario__tipo='ventas'
    ).order_by('username')

    context = {
        'ventas':       ventas,
        'fecha_desde':  request.GET.get('fecha_desde', ''),
        'fecha_hasta':  request.GET.get('fecha_hasta', ''),
        'vendedor_id':  vendedor_id,
        'vendedores':   vendedores,
        'total_ventas': ventas.count(),
        'monto_total':  ventas.aggregate(t=Sum('valor_total'))['t'] or 0,
    }

    return render(request, 'consultar_ventas.html', context)


def exportar_ventas_excel(ventas, fecha_desde, fecha_hasta):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    title_font  = Font(bold=True, size=14)

    ws.merge_cells('A1:K1')
    ws['A1'] = 'REPORTE DE VENTAS ENTREGADAS'
    ws['A1'].font      = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    if fecha_desde or fecha_hasta:
        ws.merge_cells('A2:K2')
        ws['A2'] = f"Periodo: {fecha_desde or 'Inicio'} hasta {fecha_hasta or 'Hoy'}"
        ws['A2'].alignment = Alignment(horizontal='center')
        fila_inicio = 4
    else:
        fila_inicio = 3

    headers = [
        'Fecha Envío', 'Venta #', 'Cliente', 'Vendedor', 'Producto',
        'Cantidad', 'Precio Compra Unit.', 'Precio Venta Unit.',
        'Costo Total', 'Venta Total', 'Ganancia'
    ]
    for col, header in enumerate(headers, start=1):
        cell           = ws.cell(row=fila_inicio, column=col)
        cell.value     = header
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center')

    fila           = fila_inicio + 1
    total_costo    = 0
    total_venta    = 0
    total_ganancia = 0

    for venta in ventas:
        for detalle in venta.detalles.all():
            costo_unit  = detalle.producto.valor_compra or 0
            precio_unit = detalle.precio_unitario
            cantidad    = detalle.cantidad
            costo_total = costo_unit * cantidad
            venta_total = detalle.subtotal
            ganancia    = venta_total - costo_total

            ws.cell(row=fila, column=1).value  = venta.fecha_envio.strftime('%d/%m/%Y %H:%M') if venta.fecha_envio else 'N/A'
            ws.cell(row=fila, column=2).value  = venta.id
            ws.cell(row=fila, column=3).value  = venta.cliente.nombre_completo
            ws.cell(row=fila, column=4).value  = venta.usuario_creador.username if venta.usuario_creador else 'N/A'
            ws.cell(row=fila, column=5).value  = detalle.producto.nombre
            ws.cell(row=fila, column=6).value  = cantidad
            ws.cell(row=fila, column=7).value  = float(costo_unit)
            ws.cell(row=fila, column=8).value  = float(precio_unit)
            ws.cell(row=fila, column=9).value  = float(costo_total)
            ws.cell(row=fila, column=10).value = float(venta_total)
            ws.cell(row=fila, column=11).value = float(ganancia)

            for col in [7, 8, 9, 10, 11]:
                ws.cell(row=fila, column=col).number_format = '$#,##0.00'

            total_costo    += costo_total
            total_venta    += venta_total
            total_ganancia += ganancia
            fila += 1

    fila += 1
    ws.merge_cells(f'A{fila}:H{fila}')
    ws.cell(row=fila, column=1).value = 'TOTALES'
    ws.cell(row=fila, column=1).font  = Font(bold=True, size=12)
    ws.cell(row=fila, column=9).value  = float(total_costo)
    ws.cell(row=fila, column=10).value = float(total_venta)
    ws.cell(row=fila, column=11).value = float(total_ganancia)

    for col in [9, 10, 11]:
        cell               = ws.cell(row=fila, column=col)
        cell.font          = Font(bold=True, size=12)
        cell.number_format = '$#,##0.00'
        cell.fill          = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for letra, ancho in zip('ABCDEFGHIJK', [18, 10, 25, 15, 30, 10, 18, 18, 15, 15, 15]):
        ws.column_dimensions[letra].width = ancho

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'ventas_entregadas_{fecha_desde or "inicio"}_{fecha_hasta or "hoy"}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ==================================
# CHOFERES
# ==================================
@login_required
def lista_choferes(request):
    choferes = Chofer.objects.all()
    return render(request, 'envios/lista_choferes.html', {'choferes': choferes})


@login_required
def crear_chofer(request):
    if request.method == 'POST':
        nombre   = request.POST.get('nombre_completo')
        telefono = request.POST.get('telefono')
        vehiculo = request.POST.get('vehiculo')
        pin      = request.POST.get('pin')
        notas    = request.POST.get('notas')

        if nombre and telefono and vehiculo and pin:
            Chofer.objects.create(
                nombre_completo=nombre,
                telefono=telefono,
                vehiculo=vehiculo,
                pin=pin,
                notas=notas
            )
            messages.success(request, 'Chofer creado correctamente')
            return redirect('lista_choferes')
        else:
            messages.error(request, 'Complete todos los campos obligatorios')

    return render(request, 'envios/crear_chofer.html')


@login_required
def editar_chofer(request, chofer_id):
    chofer = get_object_or_404(Chofer, id=chofer_id)

    if request.method == 'POST':
        chofer.nombre_completo = request.POST.get('nombre_completo')
        chofer.telefono        = request.POST.get('telefono')
        chofer.vehiculo        = request.POST.get('vehiculo')
        chofer.notas           = request.POST.get('notas')
        chofer.activo          = request.POST.get('activo') == 'on'
        chofer.save()

        messages.success(request, f'Chofer "{chofer.nombre_completo}" actualizado')
        return redirect('lista_choferes')

    return render(request, 'envios/editar_chofer.html', {'chofer': chofer})


@login_required
def asignar_chofer_venta(request, venta_id):
    venta = get_object_or_404(Ventas, id=venta_id, estado__in=['pendiente', 'confirmada'])

    if request.method == 'POST':
        chofer_id = request.POST.get('chofer')

        if not chofer_id:
            messages.error(request, "Debes seleccionar un chofer")
            return redirect('detalle_venta', venta_id=venta.id)

        try:
            chofer = Chofer.objects.get(id=chofer_id, activo=True)
        except Chofer.DoesNotExist:
            messages.error(request, "❌ El chofer no existe o no está activo")
            return redirect('detalle_venta', venta_id=venta.id)

        # Descontar stock si la venta está pendiente
        if venta.estado == 'pendiente':
            try:
                detalles = venta.detalles.select_related('producto').all()

                for detalle in detalles:
                    if detalle.producto.cantidad < detalle.cantidad:
                        messages.error(
                            request,
                            f"❌ Stock insuficiente de '{detalle.producto.nombre}'. "
                            f"Disponible: {detalle.producto.cantidad}, Necesario: {detalle.cantidad}"
                        )
                        return redirect('detalle_venta', venta_id=venta.id)

                for detalle in detalles:
                    detalle.producto.cantidad = F('cantidad') - detalle.cantidad
                    detalle.producto.save()
                    detalle.producto.refresh_from_db()

                messages.success(request, "✅ Stock descontado correctamente")

            except Exception as e:
                messages.error(request, f"❌ Error al descontar stock: {str(e)}")
                return redirect('detalle_venta', venta_id=venta.id)

        venta.chofer = chofer
        venta.estado = 'confirmada'

        fecha = request.POST.get('fecha_envio_programada')
        hora  = request.POST.get('hora_envio_programada')

        if fecha:
            venta.fecha_envio_programada = fecha
        if hora:
            venta.hora_envio_programada = hora

        venta.save()

        messages.success(request, f"✅ Venta #{venta.id} confirmada y asignada a {chofer.nombre_completo}")
        return redirect('detalle_venta', venta_id=venta.id)

    choferes = Chofer.objects.filter(activo=True)
    return render(request, 'envios/asignar_chofer.html', {
        'venta': venta,
        'choferes': choferes,
    })


@login_required
def asignar_envios_pendientes(request):
    cliente_id = request.GET.get('cliente')
    fecha      = request.GET.get('fecha')

    ventas = Ventas.objects.filter(
        estado__in=['pendiente', 'confirmada']
    ).select_related('cliente', 'chofer').prefetch_related('detalles__producto')

    if cliente_id:
        ventas = ventas.filter(cliente_id=cliente_id)
    if fecha:
        ventas = ventas.filter(fecha_envio_programada=fecha)  # ← corregido

    ventas = ventas.order_by('estado', '-fecha_creacion')

    context = {
        'ventas_pendientes':  ventas.filter(estado='pendiente'),
        'ventas_confirmadas': ventas.filter(estado='confirmada'),
        'clientes':   Cliente.objects.all().order_by('nombre_completo'),
        'choferes':   Chofer.objects.filter(activo=True).order_by('nombre_completo'),
        'cliente_id': cliente_id,
        'fecha':      fecha,
    }

    return render(request, 'envios/asignar_envios_pendientes.html', context)

# ==================================
# PANEL DE CHOFERES
# ==================================
@login_required
def acceso_chofer(request):
    choferes = Chofer.objects.filter(activo=True).order_by('nombre_completo')

    if request.method == 'POST':
        chofer_id = request.POST.get('chofer_id')
        pin       = request.POST.get('pin')

        chofer = get_object_or_404(Chofer, id=chofer_id, activo=True)

        if pin == chofer.pin:
            request.session['chofer_id'] = chofer.id
            return redirect('panel_chofer')
        else:
            messages.error(request, 'PIN incorrecto')

    return render(request, 'choferes/acceso_chofer.html', {'choferes': choferes})


def chofer_cerrar_sesion(request):
    request.session.pop('chofer_id', None)
    return redirect('acceso_chofer')


def panel_chofer(request):
    chofer_id = request.session.get('chofer_id')
    if not chofer_id:
        return redirect('acceso_chofer')

    chofer = get_object_or_404(Chofer, id=chofer_id, activo=True)

    fecha_filtro = request.GET.get('fecha') or date.today().strftime('%Y-%m-%d')
    fecha_parsed = None

    if fecha_filtro:
        try:
            fecha_parsed = datetime.strptime(fecha_filtro, '%Y-%m-%d').date()
        except ValueError:
            fecha_filtro = None

    if fecha_parsed:
        ventas_qs = Ventas.objects.filter(
            chofer=chofer,
            estado__in=['confirmada', 'enviada'],
            fecha_envio_programada=fecha_parsed  # ← filtra por fecha de envío estimado
        ).select_related('cliente').prefetch_related('detalles__producto').order_by(
            'hora_envio_programada'
        )
    else:
        ventas_qs = Ventas.objects.none()

    return render(request, 'choferes/panel_chofer.html', {
        'chofer':           chofer,
        'ventas_asignadas': ventas_qs,
        'fecha_filtro':     fecha_filtro,
    })

@login_required
def chofer_detalle_venta_confirmada(request, venta_id):
    chofer_id = request.session.get('chofer_id')
    if not chofer_id:
        return redirect('acceso_chofer')

    venta    = get_object_or_404(Ventas, id=venta_id, chofer_id=chofer_id)
    detalles = venta.detalles.select_related('producto').all()

    if request.method == 'POST':

        # Solo transiciones permitidas para el chofer
        TRANSICIONES_CHOFER = {
            'confirmada': 'enviada',
            'enviada':    'entregada',
        }

        nuevo_estado_esperado = TRANSICIONES_CHOFER.get(venta.estado)

        if not nuevo_estado_esperado:
            messages.error(request, '❌ Esta venta no se puede modificar.')
            return redirect('chofer_detalle_venta_confirmada', venta_id=venta_id)

        nuevo_estado = request.POST.get('estado')

        if nuevo_estado != nuevo_estado_esperado:
            messages.error(request, '❌ Transición de estado no permitida.')
            return redirect('chofer_detalle_venta_confirmada', venta_id=venta_id)

        # Método de pago
        metodo_pago    = request.POST.get('metodo_pago')
        metodos_validos = [m[0] for m in Ventas.METODO_PAGO_CHOICES]
        if metodo_pago and metodo_pago in metodos_validos:
            venta.metodo_pago = metodo_pago

        # Estado
        estado_anterior = venta.estado
        venta.estado    = nuevo_estado

        # Fecha de envío real
        if nuevo_estado == 'entregada' and estado_anterior != 'entregada' and not venta.fecha_envio:
            venta.fecha_envio = timezone.now()

        # Notas
        notas = request.POST.get('notas_adicionales')
        if notas:
            timestamp  = timezone.now().strftime('%d/%m/%Y %H:%M')
            venta.notas = (venta.notas + f"\n[{timestamp}] {notas}") if venta.notas else f"[{timestamp}] {notas}"

        venta.save()
        messages.success(request, f'✅ Estado actualizado: {venta.get_estado_display()}')
        return redirect('chofer_detalle_venta_confirmada', venta_id=venta.id)

    return render(request, 'choferes/detalle_venta_confirmada.html', {
        'venta':        venta,
        'detalles':     detalles,
        'metodos_pago': Ventas.METODO_PAGO_CHOICES,
    })
# ==================================
# IMÁGENES
# ==================================
APPS_SCRIPT_URL = "https://script.google.com/macros/s/AKfycby29kN-x1ilbQU0yEbdFTyV0E_Eyw92uzTEWyVKP-Zq0qFAoHXIdoorf_IYeSnbBHw/exec"


def subir_imagen(request):
    productos = Producto.objects.all()

    if request.method == "POST":
        producto_id = request.POST.get("producto")
        imagen = request.FILES.get("imagen")

        if producto_id and imagen:
            producto = Producto.objects.get(id=producto_id)
            imagen_base64 = base64.b64encode(imagen.read()).decode()

            r = requests.post(APPS_SCRIPT_URL, data={
                "nombre": f"producto_{producto.id}",
                "imagen": imagen_base64
            })

            respuesta = r.json()

            if respuesta.get("status") == "ok":
                # BORRAR imagen anterior si existe
                ImagenProducto.objects.filter(producto=producto).delete()
                
                # Crear nueva
                ImagenProducto.objects.create(
                    producto=producto,
                    ruta=respuesta["url"]
                )

    return render(request, "productos/subir_imagen.html", {
        "productos": productos
    })


def api_productos(request):
    lista = []
    for p in Producto.objects.all().order_by('nombre'):
        img = p.imagenproducto_set.first()
        imagen_url = img.ruta if img else None
        lista.append({
            'id':     p.id,
            'nombre': p.nombre,
            'valor':  float(p.valor),
            'imagen': imagen_url,
            'tipo':   p.tipo_id,  # ← esto es lo que faltaba
        })
    return JsonResponse(lista, safe=False)


def api_tipos(request):
    tipos = TipoProducto.objects.all().order_by('nombre')
    lista = [{'id': t.id, 'nombre': t.nombre} for t in tipos]
    return JsonResponse(lista, safe=False)


#PRUEBA

@login_required
def exportar_productos(request):
    """Descarga Excel con todos los productos actuales."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    fill_hdr  = PatternFill("solid", fgColor="0F3460")
    fill_prop = PatternFill("solid", fgColor="1B4332")
    fill_verde = PatternFill("solid", fgColor="F0FFF4")
    font_hdr  = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    font_dato = Font(color="212529", size=9, name="Arial")
    font_prop = Font(color="D8F3DC", size=9, name="Arial", italic=True)
    font_muted = Font(color="6C757D", size=9, name="Arial")
    centro = Alignment(horizontal="center", vertical="center")
    izq    = Alignment(horizontal="left",   vertical="center")
    der    = Alignment(horizontal="right",  vertical="center")

    # Fila 1: título
    ws.merge_cells("A1:I1")
    ws["A1"].value     = f"  PRODUCTOS — Exportado el {timezone.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A1"].font      = Font(bold=True, color="FFFFFF", size=12, name="Arial")
    ws["A1"].fill      = PatternFill("solid", fgColor="1A1A2E")
    ws["A1"].alignment = izq
    ws.row_dimensions[1].height = 30

    # Fila 2: encabezados
    HEADERS = [
        ("ID",               "A",  8,  fill_hdr,  centro),
        ("NOMBRE",           "B",  40, fill_hdr,  izq),
        ("TIPO",             "C",  25, fill_hdr,  izq),
        ("CANTIDAD",         "D",  12, fill_hdr,  centro),
        ("VALOR_VENTA",      "E",  15, fill_hdr,  der),
        ("VALOR_COMPRA",     "F",  15, fill_hdr,  der),
        ("GANANCIA_UNIT.$",  "G",  16, fill_prop, der),
        ("MARGEN_%",         "H",  12, fill_prop, centro),
        ("UMBRAL_ALERTA",    "I",  14, fill_hdr,  centro),
    ]
    ws.row_dimensions[2].height = 26
    for titulo, col, ancho, fill, alin in HEADERS:
        c = ws[f"{col}2"]
        c.value = titulo; c.font = font_hdr; c.fill = fill; c.alignment = alin
        ws.column_dimensions[col].width = ancho

    # Datos
    productos = Producto.objects.select_related('tipo').all().order_by('tipo__nombre', 'nombre')
    for idx, p in enumerate(productos, start=3):
        bg = "FFFFFF" if idx % 2 == 0 else "F8F9FA"
        ff = PatternFill("solid", fgColor=bg)
        ws.row_dimensions[idx].height = 16

        def c(col, val, font, fill, alin, fmt=None):
            cell = ws[f"{col}{idx}"]
            cell.value = val; cell.font = font; cell.fill = fill; cell.alignment = alin
            if fmt: cell.number_format = fmt

        c("A", p.pk,                         font_muted, ff,          centro)
        c("B", p.nombre,                     font_dato,  ff,          izq)
        c("C", p.tipo.nombre,                font_muted, ff,          izq)
        c("D", p.cantidad,                   font_dato,  ff,          centro, "0")
        c("E", float(p.valor),               font_dato,  ff,          der,    "$#,##0.00")
        c("F", float(p.valor_compra),        font_dato,  ff,          der,    "$#,##0.00")
        c("G", round(p.ganancia_unitaria,2), font_prop,  fill_verde,  der,    "$#,##0.00")
        c("H", round(p.margen_porcentaje,2), font_prop,  fill_verde,  centro, '0.00"%"')
        c("I", p.umbral_alerta,              font_dato,  ff,          centro, "0")

    ws.freeze_panes = "A3"

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="productos_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    wb.save(response)
    return response


@login_required
def descargar_template(request):
    """Descarga Excel vacío con dropdown de tipos para completar e importar."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()

    # ── Hoja 1: IMPORTAR ────────────────────────────────────
    ws = wb.active
    ws.title = "IMPORTAR"

    fill_hdr  = PatternFill("solid", fgColor="0F3460")
    fill_tipo = PatternFill("solid", fgColor="1B4332")
    font_hdr  = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    font_info = Font(bold=True, color="856404", size=9, name="Arial")
    font_normal = Font(color="212529", size=9, name="Arial")
    centro = Alignment(horizontal="center", vertical="center")
    izq    = Alignment(horizontal="left",   vertical="center")

    ws.merge_cells("A1:F1")
    ws["A1"].value     = "⚠️  Completá desde la fila 3. No modifiques los encabezados. El TIPO debe coincidir exactamente con la hoja TIPOS."
    ws["A1"].font      = font_info
    ws["A1"].fill      = PatternFill("solid", fgColor="FFF3CD")
    ws["A1"].alignment = izq
    ws.row_dimensions[1].height = 22

    COLS = [
        ("NOMBRE *",      "A", 38, fill_hdr),
        ("TIPO *",        "B", 28, fill_tipo),
        ("CANTIDAD",      "C", 12, fill_hdr),
        ("VALOR_VENTA",   "D", 15, fill_hdr),
        ("VALOR_COMPRA",  "E", 15, fill_hdr),
        ("UMBRAL_ALERTA", "F", 15, fill_hdr),
    ]
    ws.row_dimensions[2].height = 24
    for titulo, col, ancho, fill in COLS:
        c = ws[f"{col}2"]
        c.value = titulo; c.font = font_hdr; c.fill = fill; c.alignment = centro
        ws.column_dimensions[col].width = ancho

    for r in range(3, 103):
        bg = "FFFFFF" if r % 2 == 0 else "F8F9FA"
        ff = PatternFill("solid", fgColor=bg)
        for col in "ABCDEF":
            cell = ws[f"{col}{r}"]
            cell.fill = ff; cell.font = font_normal
            cell.alignment = izq if col == "A" else centro
        ws[f"D{r}"].number_format = '$#,##0.00'
        ws[f"E{r}"].number_format = '$#,##0.00'
        ws.row_dimensions[r].height = 16

    # Dropdown de tipos
    tipos = TipoProducto.objects.all().order_by('nombre')
    if tipos.exists():
        dv = DataValidation(
            type="list",
            formula1=f"TIPOS!$A$3:$A${tipos.count() + 2}",
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="Tipo inválido",
            error="Elegí un tipo de la lista desplegable.",
        )
        ws.add_data_validation(dv)
        dv.sqref = "B3:B102"

    ws.freeze_panes = "A3"

    # ── Hoja 2: TIPOS ────────────────────────────────────────
    ws2 = wb.create_sheet("TIPOS")
    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 10

    ws2.merge_cells("A1:B1")
    ws2["A1"].value     = "Tipos de producto — NO MODIFICAR esta hoja"
    ws2["A1"].font      = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    ws2["A1"].fill      = PatternFill("solid", fgColor="E94560")
    ws2["A1"].alignment = centro
    ws2.row_dimensions[1].height = 22

    for col, txt in [("A", "NOMBRE"), ("B", "ID")]:
        ws2[f"{col}2"].value = txt
        ws2[f"{col}2"].font  = Font(bold=True, color="FFFFFF", name="Arial")
        ws2[f"{col}2"].fill  = PatternFill("solid", fgColor="0F3460")
        ws2[f"{col}2"].alignment = centro

    for i, tipo in enumerate(tipos, start=3):
        ws2[f"A{i}"].value = tipo.nombre
        ws2[f"A{i}"].font  = Font(name="Arial", size=9)
        ws2[f"A{i}"].alignment = izq
        ws2[f"B{i}"].value = tipo.pk
        ws2[f"B{i}"].font  = Font(name="Arial", size=9, color="6C757D")
        ws2[f"B{i}"].alignment = centro

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="template_importar_productos.xlsx"'
    wb.save(response)
    return response


@login_required
def importar_productos(request):
    """Importa desde el template. Actualiza si existe, crea si no."""
    if request.method == 'GET':
        return render(request, 'productos/importar_productos.html')

    archivo = request.FILES.get('archivo')
    if not archivo or not archivo.name.endswith('.xlsx'):
        messages.error(request, '❌ Seleccioná un archivo .xlsx válido.')
        return render(request, 'productos/importar_productos.html')

    try:
        from openpyxl import load_workbook
        wb = load_workbook(archivo, data_only=True)
        ws = wb["IMPORTAR"] if "IMPORTAR" in wb.sheetnames else wb.active

        creados = actualizados = 0
        errores = []

        for row in ws.iter_rows(min_row=3, values_only=True):
            nombre = str(row[0] or '').strip() if row[0] else ''
            if not nombre:
                continue
            try:
                tipo_nombre  = str(row[1] or 'GENERAL').strip()
                cantidad     = int(float(row[2] or 0))
                valor        = Decimal(str(row[3] or 0))
                valor_compra = Decimal(str(row[4] or 0))
                umbral       = int(float(row[5] or 5))

                tipo, _ = TipoProducto.objects.get_or_create(nombre=tipo_nombre)
                producto = Producto.objects.filter(nombre__iexact=nombre).first()

                if producto:
                    producto.tipo = tipo; producto.cantidad = cantidad
                    producto.valor = valor; producto.valor_compra = valor_compra
                    producto.umbral_alerta = umbral; producto.save()
                    actualizados += 1
                else:
                    Producto.objects.create(
                        nombre=nombre, tipo=tipo, cantidad=cantidad,
                        valor=valor, valor_compra=valor_compra, umbral_alerta=umbral,
                    )
                    creados += 1
            except Exception as e:
                errores.append(f"'{nombre}': {e}")

        msg = f'✅ Creados: {creados} | Actualizados: {actualizados}'
        if errores:
            msg += f' | ⚠️ {len(errores)} errores'
            for err in errores[:3]:
                messages.warning(request, err)
        messages.success(request, msg)

    except Exception as e:
        messages.error(request, f'❌ Error al leer el archivo: {e}')

    return redirect('lista_productos')